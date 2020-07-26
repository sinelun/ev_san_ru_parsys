from django.contrib import admin
from ..models import Parsing, File, FileData
from django.contrib import messages
from openpyxl import load_workbook
import os


def parse_files(modeladmin, request, queryset):
    """ Парсинг прайсов из файлов в формате xlsx.
    """
    parsing = Parsing(type='files')
    parsing.save()
    for file in queryset:
        if file.parsed:
            modeladmin.message_user(request, f'Файл {file} уже был обработан ранее.', messages.WARNING)
            continue
        try:
            wb = load_workbook(file.file.path, read_only=True)
        except:
            modeladmin.message_user(request, f'Невозможно прочитать файл {file}', messages.ERROR)
            continue
        for ws in wb:
            bc = file.brand_cell.split('/')
            brand = str(ws.cell(row=int(bc[0].strip()), column=int(bc[1].strip())).value)
            if not brand:
                continue
            for x in range(1, file.max_rows):
                try:
                    v = str(ws.cell(row=x, column=file.col_num_rows).value)
                    sku = str(ws.cell(row=x, column=file.col_sku).value)
                    if v.isdigit() and sku:
                        data = {
                            'parsing': parsing,
                            'file': file,
                            'sheet': str(ws.title),
                            'brand': brand,
                            'row': int(v),
                            'sku': sku,
                            'name': str(ws.cell(row=x, column=file.col_name).value),
                            'price': str(ws.cell(row=x, column=file.col_price).value),
                        }
                        print(data)
                        FileData.objects.create(**data)
                except:
                    modeladmin.message_user(request,
                                            f'Ошибка данных в cтроке {str(v)} в файле {file.file.path}',
                                            messages.WARNING)
        parsing.completed = True
        parsing.save()
        file.parsed = True
        file.save()
        modeladmin.message_user(request, f'Файл {file} успешно обработан.', messages.SUCCESS)


parse_files.short_description = 'Парсить выбранные файлы'


@admin.register(File)
class FileAdmin(admin.ModelAdmin):
    list_display = ['file', 'uploaded', 'parsed']
    list_display_links = None
    list_editable = ['file']
    fields = ['file', 'brand_cell', 'col_num_rows', 'col_sku', 'col_name', 'col_price', 'max_rows']

    actions = [parse_files]

    def delete_queryset(self, request, queryset):
        for file in queryset:
            print(file.file.path)
            print(file.file.path.rpartition('\\')[0])
            try:
                os.remove(file.file.path)
                os.rmdir(file.file.path.rpartition('/')[0])
                os.rmdir(file.file.path.rpartition('\\')[0])
            except:
                pass  # may be todo
        super().delete_queryset(request, queryset)


@admin.register(FileData)
class FileDataAdmin(admin.ModelAdmin):
    list_display = ['parsing', 'brand', 'row', 'sku', 'name', 'price', 'created']
    list_filter = ['parsing', 'brand']
    list_display_links = None
    search_fields = ('sku', 'name')

    def has_add_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        return False
