from .models import Parsing, File, FileData
from openpyxl import load_workbook
from django.contrib import messages
import time

import logging
logger = logging.getLogger('debug')  # rename to do not log


class ParseFiles:
    """ Парсинг файлов с прайсами в отложенном режиме (cron)
    """
    def __init__(self, parsing=None, modeladmin=None, request=None, queryset=None):

        logger.info(f'The parameter parsing = {parsing}')

        self.parsing = parsing

        if not self.parsing:
            self.parsing = Parsing(type='files')
            self.parsing.save()
            logger.info(f'The new parsing -- {self.parsing} --  is starting now. --')

        logger.info(f'Parsing -- {self.parsing} --  was started -- {self.parsing.start} --')

        # Если парсинг завершён, ничего не делать
        self.do_nothing = self.parsing.completed
        logger.info(f'Parsing completed = -- {self.parsing.completed} -- | do_nothing = -- {self.do_nothing} --')

        if not self.do_nothing:
            # Файлы для парсинга
            self.queryset = queryset if queryset else File.objects.filter(parsing=self.parsing, parsed=False)

            logger.info(f'Files for parsing: queryset = -- {self.queryset} -- ')

            # Для сообщений в админ. панели
            self.modeladmin = modeladmin
            self.request = request

    def parse(self):
        if self.do_nothing:
            logger.info(f'The current parsing -- {self.parsing} -- has already completed and script will do nothing')
            return
        for file in self.queryset:
            if file.parsed:
                message = f'Файл {file} уже был обработан ранее.'
                self._admin_message(message, messages.WARNING)
                logger.info(message)
                continue
            try:
                wb = load_workbook(file.file.path, read_only=True)
                logger.info(f'Was the workbook loaded from -- {file.file.path} -- ? - -- {bool(wb)} --')
            except:
                message = f'Невозможно прочитать файл {file}'
                self._admin_message(message, messages.ERROR)
                logger.info(message)
                continue
            # Сопоставляем файл с парсингом перед обработкой
            logger.info(f'The file.parsing was -- {file.parsing} -- before.')
            file.parsing = self.parsing
            file.save()
            logger.info(f'The file.parsing is -- {file.parsing} -- now.')
            # Перебираем все excel листы
            logger.info(f'last_parsed_sheet: -- {file.last_parsed_sheet} -- .')
            for i, ws in enumerate(wb, start=1):
                logger.info(f'The current worksheet number is {i}.')
                # Пропускам листы до последнего не обработанного
                if i <= file.last_parsed_sheet:
                    continue
                # Парсинг текущего листа происходит и возвращает необходимость продолжения парсинга того же листа
                continue_this_sheet = self._parse_worksheet(file, i, ws)
                # Если парсинг того же листа следует продолжить, то не надо менять номер последнего обработанного листа
                if continue_this_sheet:
                    return
                # Если лист не последний, парсинг останавливается, чтобы быть продолженным через cron,...
                if i < len(wb.sheetnames):
                    # Сохраняем номер последнего обработанного листа и прекращаем работу
                    file.last_parsed_sheet = i
                    file.save()
                    logger.info(f'The new last_parsed_sheet = -- {file.last_parsed_sheet} --.')
                    return
                # ... иначе ставим метку, что файл обработан
                else:
                    file.parsed = True
                    file.save()
                    logger.info(f'File {file} has parsed')
            # По окончании делаем парсинг завершёным
            self.parsing.completed = True
            self.parsing.save()
            logger.info(f'Parsing {self.parsing} has completed')

    def _parse_worksheet(self, file, index, ws):
        """ Парсинг части рабочего листа excel от точки останова (file.last_parsed_row + 1)
            до конца или до плюс number_of_next_rows строк
        """
        logger.info(f'Parsing of worksheet -- {index} -- started.')
        # Ячейка с брендом,...
        bc = file.brand_cell.split('/')
        brand = str(ws.cell(row=int(bc[0].strip()), column=int(bc[1].strip())).value)
        logger.info(f'The brand name of the current worksheet is -- {brand} --.')
        # ... если пуста, значит лист не содержит данных прайса
        if not brand:
            logger.info(f'The current worksheet is skipped because the brand is empty.')
            return
        # (todo settings) следующая порция строк
        number_of_next_rows = 200;
        logger.info(f'The number of the next rows has set to -- {number_of_next_rows} -- .')
        start_row = file.last_parsed_row + 1
        # Если параметр максимального количства строк меньше, чем номер предполагаемой последней строки...
        if file.max_rows < start_row + number_of_next_rows:
            # ... максимум - параметр максимума строк
            max_row = file.max_rows
            # Сброс последней обработанной строки для следующего листа, так как в текущем будет достинут максимум строк.
            last_parsed_row = 0
            # Текущему листу не требуется продолжение обработки
            requires_continuation = False
        else:
            # ... максимум = старт + следующая порция строк
            max_row = start_row + number_of_next_rows - 1
            # Обновление номера последней строки.
            last_parsed_row = max_row
            # Текущему листу требуется продолжение обработки
            requires_continuation = True

        logger.info(f'The new value of file.last_parsed_row = -- {file.last_parsed_row} --.')
        logger.info(f'The current worksheet will be parsed from row -- {start_row} --  to -- {max_row} --.')

        # Для проверки на пустоту (вакуум): нет смысла проходить по пустым строкам, следует перейти к следующему листу
        void = True
        start_time = time.time()
        for x in range(start_row, max_row+1):
            try:
                v = str(ws.cell(row=x, column=file.col_num_rows).value)
                sku = str(ws.cell(row=x, column=file.col_sku).value)
                if v.isdigit() and sku:
                    void = False
                    data = {
                        # 'file': file,
                        # 'sheet': str(ws.title),
                        'brand': brand,
                        # 'row': int(v),
                        'sku': sku,
                        'name': str(ws.cell(row=x, column=file.col_name).value),
                        'price': str(ws.cell(row=x, column=file.col_price).value),
                    }
                    print(data)
                    record, created = FileData.objects.update_or_create(
                        file=file, sheet=str(ws.title), row=int(v), defaults=data)
                    logger.info(f'Created = {created} '
                                f'-- {record.file} -- {record.sheet} -- {record.row} -- {record.sku} -- {record.price}')
            except:
                message = f'Ошибка данных в cтроке {str(v)} на листе {str(ws.title)} в файле {file.file.path}'
                self._admin_message(message, messages.WARNING)
                logger.warning(message)

        execution_time = time.time() - start_time

        if void:
            logger.info(f'Void detected.')
            last_parsed_row = 0
            requires_continuation = False

        # Номер последней обработанной строки останется прежним в случае сбоя, что даст повторение предыдущего цикла
        file.last_parsed_row = last_parsed_row
        file.save()

        logger.info(f'Parsing session completed in {execution_time} sec. '
                    f'Does it require continuation? -- {requires_continuation} --')

        return requires_continuation;

    def _admin_message(self, message, level):
        if not self.modeladmin or not self.request:
            return
        self.modeladmin.message_user(self.request, message, level)
